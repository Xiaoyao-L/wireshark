import pyshark
from collections import Counter
import matplotlib.pyplot as plt
import numpy as np
import datetime
from openpyxl import Workbook


# transmission address
def getaddress_ta(packets):
    addr_list = []
    addr_solved = []
    for pkt in packets:
        addr_list.append(pkt.wlan.ta)

    addr_count = Counter(addr_list)

    for k, v in addr_count.items():
        com = ''
        if k is not None:
            fp = open('manuf', 'r', encoding='utf-8')
            lines = fp.readlines()
            for line in lines:
                if line.find(k.upper()[0:8]) != -1:
                    com = line[9:].strip()
                    break
            fp.close()
        addr_solved.append([com.strip().replace('\t', ' '), k, v])

    return addr_solved


# destination address
def getaddress_da(packets):
    addr_list = []
    addr_solved = []
    for pkt in packets:
        addr_list.append(pkt.wlan.da)

    addr_count = Counter(addr_list)

    for k, v in addr_count.items():
        com = ''
        if k is not None:
            fp = open('manuf', 'r', encoding='utf-8')
            lines = fp.readlines()
            for line in lines:
                if line.find(k.upper()[0:8]) != -1:
                    com = line[9:].strip()
                    break
            fp.close()
        addr_solved.append([com.strip().replace('\t', ' '), k, v])

    return addr_solved


# station address
def getaddress_sta(packets_1, packets_2):
    addr_list = []
    addr_solved = []
    for pkt in packets_1:
        addr_list.append(pkt.wlan.ta)

    for pkt in packets_2:
        addr_list.append(pkt.wlan.da)

    addr_count = Counter(addr_list)

    for k, v in addr_count.items():
        com = ''
        if k is not None:
            fp = open('manuf', 'r', encoding='utf-8')
            lines = fp.readlines()
            for line in lines:
                if line.find(k.upper()[0:8]) != -1:
                    com = line[9:].strip()
                    break
            fp.close()
        addr_solved.append([com.strip().replace('\t', ' '), k, v])

    print(addr_solved)
    return addr_solved


# get vendor type
def getvendorname(address):
    addr_solved = []
    for k in address:
        flag = 0
        if k is not None:
            fp = open('manuf', 'r', encoding='utf-8')
            lines = fp.readlines()
            for line in lines:
                if line.find(k.upper()[0:8]) != -1:
                    k = line[9:].strip()
                    flag = 1
                    break
            fp.close()
        if flag == 0:
            k = 'Not Found'
        addr_solved.append(k.strip().replace('\t', ' '))
    return addr_solved


# vendors' broadcast
def test(packets):
    addr_list = []
    addr_solved = []
    for pkt in packets:
        addr_list.append(pkt.wlan.da)

    for k in addr_list:
        if k is not None:
            fp = open('manuf', 'r', encoding='utf-8')
            lines = fp.readlines()
            for line in lines:
                if line.find(k.upper()[0:8]) != -1:
                    k = line[9:].strip()
                    break
            fp.close()
        addr_solved.append(k.strip().replace('\t', ' '))
    addr_solved_count = Counter(addr_solved)
    return addr_solved_count

# export to excel
def export_(data, filename):
    out = Workbook()
    out_ = out.create_sheet(index=0)
    index = 0
    for key, value in data.items():
        index += 1
        out_.cell(index, 1).value = key
        out_.cell(index, 2).value = value
    out.save(filename)


# pie
def fig_pie(labels, sizes, title):
    plt.figure(figsize=(10, 10))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', shadow=False, startangle=90)
    plt.axis('equal')
    plt.title(title)
    # plt.tight_layout()
    plt.savefig(title + '.png', bbox_inches='tight')


# bar
def fig_bar(labels, sizes, title):
    print(labels)
    print(sizes)
    bar_width = 0.3
    plt.figure(figsize=(10, 10))
    plt.barh(y=range(len(labels)), width=sizes, color='#6699CC')

    for y, x in enumerate(sizes):
        plt.text(x + 20, y - bar_width / 2, '%s' % x, ha='center', va='bottom', fontsize=20)

    plt.yticks(np.arange(len(labels)) + bar_width / 2, labels, fontsize=20)

    plt.title(title, fontsize=20)
    plt.tight_layout()
    plt.savefig(title + '_bar' + '.png', bbox_inches='tight')


# number <= 3, transfer to others
def toOthers(data):
    new_data = dict()
    i = 0
    for key, value in data.items():
        if value is not None:
            if value <= 3:
                i = i + 1
            if value > 3:
                new_data[key] = value
    new_data['Others'] = i
    print(new_data)
    return new_data


# collect station information
def station(filename):
    start_time = datetime.datetime.now()
    cap_sta_1 = pyshark.FileCapture(filename, display_filter='wlan.fc.type_subtype eq 4')
    cap_sta_2 = pyshark.FileCapture(filename, display_filter='wlan.fc.type_subtype eq 5')
    addr_sta = getaddress_sta(cap_sta_1, cap_sta_2)
    cap_sta_1.close()
    cap_sta_2.close()

    labels_sta = []

    for addr in addr_sta:
        labels_sta.append(addr[:][1].upper()[0:8])

    print(labels_sta)

    vendor_sta_count = Counter(getvendorname(labels_sta))

    vendor_sta_sorted = dict(sorted(vendor_sta_count.items(), key=lambda x: x[1], reverse=True))

    print(vendor_sta_sorted)

    export_(vendor_sta_sorted, 'station_on campus.xlsx')

    data_reform = toOthers(vendor_sta_sorted)

    fig_pie(data_reform.keys(), data_reform.values(), 'Vendors distribution of Stations on campus')

    fig_bar(list(data_reform.keys()), list(data_reform.values()), 'Vendors distribution of Stations '
                                                                  'on campus')

    # plt.show()

    print('finish station')

    end_time = datetime.datetime.now()

    print(end_time - start_time)


# collect ap information
def ap(filename, filter_='wlan.fc.type_subtype eq 8'):
    start_time = datetime.datetime.now()
    cap_ap = pyshark.FileCapture(filename, display_filter=filter_)
    addr_ap = getaddress_ta(cap_ap)
    # addr_sta_count = test(cap_sta)
    # print(addr_sta_count)
    cap_ap.close()

    labels_ap = []

    for addr in addr_ap:
        labels_ap.append(addr[:][1].upper()[0:8])

    vendor_ap_count = Counter(getvendorname(labels_ap))

    vendor_ap_sorted = dict(sorted(vendor_ap_count.items(), key=lambda x: x[1], reverse=True))

    print(vendor_ap_sorted)

    export_(vendor_ap_sorted, 'ap_on campus.xlsx')

    data_reform = toOthers(vendor_ap_sorted)

    fig_pie(data_reform.keys(), data_reform.values(), 'Vendors distribution of Access Points on campus')

    fig_bar(list(data_reform.keys()), list(data_reform.values()), 'Vendors distribution of Access Points '
                                                                  'on campus')

    # plt.show()

    print('finish Access Point')

    end_time = datetime.datetime.now()

    print(end_time - start_time)


# main function
def analyze(filename):
    station(filename)
    ap(filename)


analyze('campus-ewi-dorm.pcap')
